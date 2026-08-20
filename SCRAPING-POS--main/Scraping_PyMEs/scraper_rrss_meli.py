import os
import re
import time
import pandas as pd
from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Cadenas y grandes marcas a excluir
CADENAS_EXCLUIDAS = [
    "falabella", "paris", "ripley", "sodimac", "easy", "lider", "walmart",
    "santa isabel", "unimarc", "alvi", "jumbo", "express", "tottus", "oxxo",
    "ok market", "mall", "plaza", "costanera", "outlet", "copec", "shell",
    "penta", "spid", "cruz verde", "ahumada", "salcobrand", "imperial",
    "construmart", "chilemat", "mts", "autoplanet", "redsalud", "integramedica",
    "clinica alemana", "clinica las condes", "clinica indisa", "bupa", "uno salud"
]

COLUMNAS_EXCEL = [
    "Nombre_Empresa",
    "Rubro",
    "Telefono_WSP",
    "Ubicacion_Zona",
    "Fuente_Deteccion",
    "Asociado_Mercado_Pago",
    "Canal_Venta",
    "Enlace_Referencia",
    "Estado_Contacto",
    "Observaciones"
]

def limpiar_texto(texto):
    if not texto:
        return ""
    texto_limpio = re.sub(r'[^\w\s\.,#\-\(\)/áéíóúÁÉÍÓÚñÑa-zA-Z0-9]', '', str(texto))
    return re.sub(r'\s+', ' ', texto_limpio).strip()

def normalizar_telefono_chile(telefono_str):
    if not telefono_str or str(telefono_str).strip() == "" or str(telefono_str).lower() == "nan":
        return None

    digitos = re.sub(r'\D', '', str(telefono_str))
    if digitos.startswith('56'):
        digitos = digitos[2:]

    # Descartar red fija (inicia con 2) o números 800 / 600
    if digitos.startswith('2') or digitos.startswith('800') or digitos.startswith('600'):
        return None

    # Validar formato celular chileno (9 dígitos comenzando con 9)
    if len(digitos) == 9 and digitos.startswith('9'):
        return f"+56 {digitos[0]} {digitos[1:5]} {digitos[5:]}"

    return None

def es_cadena_grande(nombre):
    nombre_lower = str(nombre).lower()
    return any(cadena in nombre_lower for cadena in CADENAS_EXCLUIDAS)


# =====================================================================
# MODULO: SCRAPING DE TIENDAS / VENDEDORES EN MERCADO LIBRE CHILE
# =====================================================================
def scrapear_mercadolibre_pymes(rubro, max_resultados, telefonos_existentes, nombres_existentes):
    print(f"\n🛒 [Mercado Libre] Buscando comercios de '{rubro}' en RM...")
    nuevos = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(locale="es-CL")
        page = context.new_page()

        # Búsqueda en ML Chile filtrando por RM
        url_busqueda = f"https://listado.mercadolibre.cl/{rubro.replace(' ', '-')}_Ubicacion_RM-Metropolitana"
        page.goto(url_busqueda)
        page.wait_for_timeout(3500)

        # Capturar tarjetas de productos
        items = page.locator('ol.ui-search-layout li.ui-search-layout__item').all()
        print(f"📦 Publicaciones detectadas en página: {len(items)}")

        for item in items:
            if len(nuevos) >= max_resultados:
                break

            try:
                # Extraer nombre del vendedor o tienda
                seller_locator = item.locator('.ui-search-item__group__element--seller, .ui-search-official-store-label')
                titulo_locator = item.locator('h2.ui-search-item__title')
                
                nombre_comercio = ""
                if seller_locator.count() > 0:
                    nombre_comercio = seller_locator.first.inner_text().replace("Por ", "").strip()
                elif titulo_locator.count() > 0:
                    nombre_comercio = titulo_locator.first.inner_text().strip()

                nombre_clean = limpiar_texto(nombre_comercio)
                if not nombre_clean or es_cadena_grande(nombre_clean) or nombre_clean.lower() in nombres_existentes:
                    continue

                # Obtener enlace del artículo/tienda
                link_locator = item.locator('a.ui-search-link')
                if link_locator.count() == 0:
                    continue
                item_url = link_locator.first.get_attribute("href")

                # Navegar al producto para extraer datos de contacto o tienda del vendedor
                page_prod = context.new_page()
                page_prod.goto(item_url)
                page_prod.wait_for_timeout(2000)

                body_text = page_prod.locator('body').inner_text()

                # Buscar patrones de WhatsApp
                match_tel = re.search(r'(\+?56\s?9\s?\d{4}\s?\d{4}|\b9\s?\d{4}\s?\d{4}\b)', body_text)
                tel_validado = None
                if match_tel:
                    tel_validado = normalizar_telefono_chile(match_tel.group(0))

                page_prod.close()

                if tel_validado and tel_validado not in telefonos_existentes:
                    telefonos_existentes.add(tel_validado)
                    nombres_existentes.add(nombre_clean.lower())

                    nuevos.append({
                        "Nombre_Empresa": nombre_clean,
                        "Rubro": rubro.capitalize(),
                        "Telefono_WSP": str(tel_validado),
                        "Ubicacion_Zona": "RM / Periferia",
                        "Fuente_Deteccion": "Mercado Libre",
                        "Asociado_Mercado_Pago": "No / En Revisión",
                        "Canal_Venta": "Online + Físico",
                        "Enlace_Referencia": item_url[:120],
                        "Estado_Contacto": "Pendiente",
                        "Observaciones": "Prospecto detectado en marketplace sin POS visible"
                    })
                    print(f"  ✅ [MeLi] Contacto capturado: {nombre_clean} | WSP: {tel_validado}")

            except Exception:
                continue

        browser.close()
    return nuevos


# =====================================================================
# MODULO: SCRAPING DE REDES SOCIALES (INSTAGRAM / LINKTREE)
# =====================================================================
def scrapear_instagram_bio(rubro, zona, max_resultados, telefonos_existentes, nombres_existentes):
    print(f"\n📸 [RRSS / Instagram] Buscando comercios de '{rubro}' en '{zona}'...")
    nuevos = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(locale="es-CL")
        page = context.new_page()

        # Búsqueda indexada pública en Google para perfiles de Instagram en RM
        query = f'site:instagram.com "{rubro}" "{zona}" "+569" -site:instagram.com/p/'
        url_google = f"https://www.google.com/search?q={query.replace(' ', '+')}"
        page.goto(url_google)
        page.wait_for_timeout(3000)

        resultados = page.locator('div.g').all()
        print(f"🔎 Perfiles públicos detectados: {len(resultados)}")

        for res in resultados:
            if len(nuevos) >= max_resultados:
                break

            try:
                snippet = res.inner_text()
                link_el = res.locator('a')
                if link_el.count() == 0:
                    continue

                perfil_url = link_el.first.get_attribute("href")
                if "instagram.com" not in perfil_url or "/p/" in perfil_url:
                    continue

                # Extraer título y limpiar nombre
                titulo_el = res.locator('h3')
                nombre_raw = titulo_el.first.inner_text() if titulo_el.count() > 0 else "Local Instagram"
                nombre_clean = limpiar_texto(re.sub(r'(@|\(\@.*\)|•.*|Instagram.*)', '', nombre_raw))

                if not nombre_clean or es_cadena_grande(nombre_clean) or nombre_clean.lower() in nombres_existentes:
                    continue

                # Extraer WhatsApp directo del snippet indexado
                match_tel = re.search(r'(\+?56\s?9\s?\d{4}\s?\d{4}|\b9\s?\d{4}\s?\d{4}\b)', snippet)
                tel_validado = None
                if match_tel:
                    tel_validado = normalizar_telefono_chile(match_tel.group(0))

                if tel_validado and tel_validado not in telefonos_existentes:
                    telefonos_existentes.add(tel_validado)
                    nombres_existentes.add(nombre_clean.lower())

                    nuevos.append({
                        "Nombre_Empresa": nombre_clean,
                        "Rubro": rubro.capitalize(),
                        "Telefono_WSP": str(tel_validado),
                        "Ubicacion_Zona": zona.capitalize(),
                        "Fuente_Deteccion": "Instagram / RRSS",
                        "Asociado_Mercado_Pago": "No",
                        "Canal_Venta": "Redes Sociales",
                        "Enlace_Referencia": perfil_url,
                        "Estado_Contacto": "Pendiente",
                        "Observaciones": "Extraído desde biografía pública"
                    })
                    print(f"  ✅ [RRSS] Contacto capturado: {nombre_clean} | WSP: {tel_validado}")

            except Exception:
                continue

        browser.close()
    return nuevos


# =====================================================================
# GESTOR DE BASE DE DATOS INCREMENTAL (EXCEL INDEPENDIENTE)
# =====================================================================
def guardar_base_rrss_meli(nuevos_registros, archivo_salida="Base_Prospectos_RRSS_MeLi.xlsx"):
    df_previo = pd.DataFrame(columns=COLUMNAS_EXCEL)

    # 1. Cargar y depurar base previa si existe
    if os.path.exists(archivo_salida):
        try:
            df_cargado = pd.read_excel(archivo_salida, dtype={"Telefono_WSP": str})
            for col in COLUMNAS_EXCEL:
                if col not in df_cargado.columns:
                    df_cargado[col] = ""
            df_previo = df_cargado[COLUMNAS_EXCEL].copy()
            # Conservar el último registro en caso de duplicados previos
            df_previo.drop_duplicates(subset=["Telefono_WSP"], keep="last", inplace=True)
        except Exception as e:
            print(f"⚠️ Error al leer la base previa: {e}")

    # 2. Unir registros antiguos con los nuevos
    df_nuevos = pd.DataFrame(nuevos_registros)
    df_final = pd.concat([df_previo, df_nuevos], ignore_index=True)
    
    # REGLA: Conservar el último registro si colisiona algún teléfono
    df_final.drop_duplicates(subset=["Telefono_WSP"], keep="last", inplace=True)

    # 3. Exportar con formato estilizado
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prospectos_Alternativos"

    ws.append(COLUMNAS_EXCEL)
    for row in df_final.itertuples(index=False):
        ws.append(list(row))

    header_fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid") # Azul oscuro
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )

    ws.row_dimensions[1].height = 28
    for col_idx in range(1, len(COLUMNAS_EXCEL) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for r_idx in range(2, len(df_final) + 2):
        ws.row_dimensions[r_idx].height = 22
        for c_idx in range(1, len(COLUMNAS_EXCEL) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = data_font
            cell.border = thin_border
            col_name = COLUMNAS_EXCEL[c_idx - 1]
            if col_name == "Telefono_WSP":
                cell.number_format = '@'
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name in ["Ubicacion_Zona", "Fuente_Deteccion", "Asociado_Mercado_Pago", "Estado_Contacto"]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Autoajuste de anchos de columna
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    ws.auto_filter.ref = ws.dimensions
    wb.save(archivo_salida)
    print(f"\n🎉 Base guardada exitosamente en '{archivo_salida}'. Total de prospectos activos: {len(df_final)}")


# =====================================================================
# EJECUCIÓN PRINCIPAL
# =====================================================================
if __name__ == "__main__":
    rubros_objetivo = [
        "Ferreteria",
        "Taller mecanico",
        "Repuestos automotrices",
        "Boutique ropa",
        "Centro de estetica",
        "Distribuidora insumos"
    ]
    zonas_santiago = ["Santiago", "Providencia", "Las Condes", "Maipu", "San Bernardo", "Colina", "Lampa"]
    archivo_db = "Base_Prospectos_RRSS_MeLi.xlsx"

    telefonos_existentes = set()
    nombres_existentes = set()

    # Pre-cargar teléfonos y nombres ya existentes
    if os.path.exists(archivo_db):
        try:
            df_actual = pd.read_excel(archivo_db, dtype={"Telefono_WSP": str})
            for _, r in df_actual.iterrows():
                tel = str(r["Telefono_WSP"]).strip()
                nom = str(r["Nombre_Empresa"]).strip().lower()
                if tel and tel != "nan":
                    telefonos_existentes.add(tel)
                if nom and nom != "nan":
                    nombres_existentes.add(nom)
            print(f"📂 Base cargada con {len(telefonos_existentes)} registros existentes.")
        except Exception:
            pass

    todos_los_prospectos = []

    # 1. Prospección en Mercado Libre
    for rubro in rubros_objetivo:
        datos_meli = scrapear_mercadolibre_pymes(
            rubro=rubro,
            max_resultados=5,
            telefonos_existentes=telefonos_existentes,
            nombres_existentes=nombres_existentes
        )
        todos_los_prospectos.extend(datos_meli)

    # 2. Prospección en Redes Sociales (Instagram)
    for rubro in rubros_objetivo:
        for zona in zonas_santiago[:3]:  # Muestreo por zona
            datos_rrss = scrapear_instagram_bio(
                rubro=rubro,
                zona=zona,
                max_resultados=3,
                telefonos_existentes=telefonos_existentes,
                nombres_existentes=nombres_existentes
            )
            todos_los_prospectos.extend(datos_rrss)

    # 3. Consolidación en Excel independiente
    if todos_los_prospectos:
        guardar_base_rrss_meli(todos_los_prospectos, archivo_salida=archivo_db)
    else:
        print("\nℹ️ No se detectaron nuevos prospectos únicos en este ciclo de búsqueda.")