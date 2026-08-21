import os
import re
import time
import pandas as pd
from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CADENAS_EXCLUIDAS = [
    "falabella", "paris", "ripley", "sodimac", "easy", "lider", "walmart",
    "santa isabel", "unimarc", "alvi", "jumbo", "express", "tottus", "oxxo",
    "ok market", "mall", "plaza", "costanera", "outlet", "copec", "shell",
    "penta", "spid", "cruz verde", "ahumada", "salcobrand", "imperial",
    "construmart", "chilemat", "mts", "autoplanet", "redsalud", "integramedica",
    "clinica alemana", "clinica las condes", "clinica indisa", "bupa", "uno salud"
]

COLUMNAS_EXCEL = [
    "Nombre_Comercio",
    "Rubro",
    "Telefono_WSP",
    "Ubicacion_Comuna",
    "Direccion_Fisica",
    "Tiene_POS_Estimado",
    "Estado_Contacto",
    "Observaciones"
]

def limpiar_texto(texto):
    if not texto:
        return ""
    texto_limpio = re.sub(r'[^\w\s\.,#\-\(\)/áéíóúÁÉÍÓÚñÑa-zA-Z0-9]', '', str(texto))
    return re.sub(r'\s+', ' ', texto_limpio).strip()

def normalizar_telefono_chile(telefono_str):
    if not telefono_str or str(telefono_str).strip() == "" or str(telefono_str).lower() == "nan":
        return None

    digitos = re.sub(r'\D', '', str(telefono_str))
    if digitos.startswith('56'):
        digitos = digitos[2:]

    # Descartar red fija (inicia con 2) o números 800/600
    if digitos.startswith('2') or digitos.startswith('800') or digitos.startswith('600'):
        return None

    # Validar celular chileno (9 dígitos comenzando con 9)
    if len(digitos) == 9 and digitos.startswith('9'):
        return f"+56 {digitos[0]} {digitos[1:5]} {digitos[5:]}"

    return None

def es_cadena_grande(nombre):
    nombre_lower = str(nombre).lower()
    return any(cadena in nombre_lower for cadena in CADENAS_EXCLUIDAS)

def extraer_locales_santiago(rubro, max_resultados, telefonos_existentes, nombres_existentes):
    comuna = "Santiago"
    print(f"\n🚀 Buscando: '{rubro}' en '{comuna}'...")
    nuevos_locales = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(locale="es-CL")
        page = context.new_page()

        query_url = f"https://www.google.com/maps/search/{rubro.replace(' ', '+')}+{comuna.replace(' ', '+')}"
        page.goto(query_url)
        page.wait_for_timeout(4000)

        for _ in range(6):
            try:
                feed = page.locator('div[role="feed"]')
                if feed.count() > 0:
                    feed.first.evaluate('el => el.scrollTop += 1500')
                else:
                    page.mouse.wheel(0, 1500)
            except Exception:
                page.mouse.wheel(0, 1500)
            page.wait_for_timeout(1200)

        links = page.locator('a[href*="/maps/place/"]').all()
        print(f"🔍 Locales detectados en pantalla: {len(links)}")

        urls_visitadas = set()

        for link in links:
            href = link.get_attribute("href")
            if not href or href in urls_visitadas:
                continue
            urls_visitadas.add(href)

            nombre_local = link.get_attribute("aria-label")
            if not nombre_local or nombre_local.strip() == "":
                continue

            nombre_clean = limpiar_texto(nombre_local)

            if es_cadena_grande(nombre_clean) or nombre_clean.lower() in nombres_existentes:
                continue

            if len(nuevos_locales) >= max_resultados:
                break

            try:
                link.click()
                page.wait_for_timeout(2200)

                # Extraer Teléfono
                telefono_raw = ""
                tel_locator = page.locator('button[data-tooltip*="teléfono"], button[aria-label*="Teléfono"], button[data-item-id*="phone"]')
                if tel_locator.count() > 0:
                    telefono_raw = tel_locator.first.inner_text().strip()
                else:
                    main_container = page.locator('div[role="main"]')
                    if main_container.count() > 0:
                        body_text = main_container.inner_text()
                        match = re.search(r'(\+?56\s?9\s?\d{4}\s?\d{4}|\b9\s?\d{4}\s?\d{4}\b)', body_text)
                        if match:
                            telefono_raw = match.group(0)

                telefono_validado = normalizar_telefono_chile(telefono_raw)
                if not telefono_validado or telefono_validado in telefonos_existentes:
                    continue

                # Extraer Dirección
                direccion = "Sin Dirección"
                dir_locator = page.locator('button[data-item-id="address"], button[aria-label*="Dirección"]')
                if dir_locator.count() > 0:
                    direccion = dir_locator.first.inner_text().strip()

                dir_clean = limpiar_texto(direccion)

                telefonos_existentes.add(telefono_validado)
                nombres_existentes.add(nombre_clean.lower())

                nuevos_locales.append({
                    "Nombre_Comercio": nombre_clean,
                    "Rubro": rubro.capitalize(),
                    "Telefono_WSP": str(telefono_validado),
                    "Ubicacion_Comuna": comuna,
                    "Direccion_Fisica": dir_clean,
                    "Tiene_POS_Estimado": "Sí (Local Físico)",
                    "Estado_Contacto": "Pendiente",
                    "Observaciones": ""
                })
                print(f"  ✅ Prospecto Guardado: {nombre_clean} | WSP: {telefono_validado}")

            except Exception:
                continue

        browser.close()

    return nuevos_locales

def guardar_base_consolidada(nuevos_registros, archivo_salida="Base_MeLi.xlsx"):
    df_previo = pd.DataFrame(columns=COLUMNAS_EXCEL)

    if os.path.exists(archivo_salida):
        try:
            df_cargado = pd.read_excel(archivo_salida, dtype={"Telefono_WSP": str})
            for col in COLUMNAS_EXCEL:
                if col not in df_cargado.columns:
                    df_cargado[col] = ""
            df_previo = df_cargado[COLUMNAS_EXCEL].copy()
            df_previo.drop_duplicates(subset=["Telefono_WSP"], keep="last", inplace=True)
        except Exception as e:
            print(f"⚠️ Error al leer base previa: {e}")

    df_nuevos = pd.DataFrame(nuevos_registros)
    df_final = pd.concat([df_previo, df_nuevos], ignore_index=True)
    df_final.drop_duplicates(subset=["Telefono_WSP"], keep="last", inplace=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prospectos_Santiago"

    ws.append(COLUMNAS_EXCEL)
    for row in df_final.itertuples(index=False):
        ws.append(list(row))

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )

    ws.row_dimensions[1].height = 28
    for col_idx in range(1, len(COLUMNAS_EXCEL) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for r_idx in range(2, len(df_final) + 2):
        ws.row_dimensions[r_idx].height = 22
        for c_idx in range(1, len(COLUMNAS_EXCEL) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = data_font
            cell.border = thin_border
            col_name = COLUMNAS_EXCEL[c_idx - 1]
            if col_name == "Telefono_WSP":
                cell.number_format = '@'
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name in ["Ubicacion_Comuna", "Tiene_POS_Estimado", "Estado_Contacto"]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    ws.auto_filter.ref = ws.dimensions
    wb.save(archivo_salida)
    print(f"\n🎉 ¡Guardado exitoso en '{archivo_salida}'! Total comercios activos en Santiago: {len(df_final)}")

if __name__ == "__main__":
    rubros = [
        "Taller mecanico",
        "Ferreteria",
        "Repuestos automotrices",
        "Materiales de construccion",
        "Clinica dental",
        "Veterinaria",
        "Centro de estetica",
        "Peluqueria barberia",
        "Distribuidora insumos",
        "Optica",
        "Tienda de ropa boutique"
    ]

    archivo_base = "Base_MeLi.xlsx"
    telefonos_existentes = set()
    nombres_existentes = set()

    if os.path.exists(archivo_base):
        try:
            df_actual = pd.read_excel(archivo_base, dtype={"Telefono_WSP": str})
            for _, r in df_actual.iterrows():
                tel = str(r["Telefono_WSP"]).strip()
                nom = str(r["Nombre_Comercio"]).strip().lower()
                if tel and tel != "nan":
                    telefonos_existentes.add(tel)
                if nom and nom != "nan":
                    nombres_existentes.add(nom)
            print(f"📂 Base cargada con {len(telefonos_existentes)} comercios previos en memoria.")
        except Exception:
            pass

    todos_los_locales = []

    for rubro_target in rubros:
        locales = extraer_locales_santiago(
            rubro=rubro_target,
            max_resultados=8,
            telefonos_existentes=telefonos_existentes,
            nombres_existentes=nombres_existentes
        )
        todos_los_locales.extend(locales)

    if todos_los_locales:
        guardar_base_consolidada(todos_los_locales, archivo_salida=archivo_base)
    else:
        print("\nℹ️ No se detectaron nuevos comercios únicos en Santiago en esta pasada.")